"""Flask dashboard. http://localhost:5000"""
from __future__ import annotations

import io
import logging
import os
import subprocess
import sys
import threading
import time
from datetime import datetime

from flask import Flask, Response, jsonify, render_template, request

from config import FLASK_HOST, FLASK_PORT, MIN_FIT_SCORE
from db import get_jobs, init_db, stats, update_status

log = logging.getLogger(__name__)
app = Flask(__name__)

_apply_lock = threading.Lock()
_apply_state = {"running": False, "done": 0, "total": 0, "current": None}


def _apply_worker(job_ids: list[int]) -> None:
    from applicator import run_apply_many

    _apply_state.update(running=True, done=0, total=len(job_ids), current=None)
    try:
        run_apply_many(job_ids)
    finally:
        _apply_state.update(running=False, current=None)


def _search_worker() -> None:
    from agent import run_pipeline_once

    try:
        run_pipeline_once()
    except Exception:  # noqa: BLE001
        log.exception("manual run failed")


def _group_jobs(jobs: list[dict]) -> dict:
    auto_apply: list[dict] = []
    manual: list[dict] = []
    for j in jobs:
        s = j.get("status", "pending")
        if s in ("pending", "approved"):
            auto_apply.append(j)
        elif s in ("failed", "needs_manual"):
            manual.append(j)
    # Companies: dedup by lowercased name, list roles
    company_map: dict[str, dict] = {}
    for j in jobs:
        c = (j.get("company") or "").strip()
        if not c:
            continue
        key = c.lower()
        if key not in company_map:
            company_map[key] = {"name": c, "roles": []}
        title = (j.get("title") or "").strip()
        if title and title not in company_map[key]["roles"]:
            company_map[key]["roles"].append(title)
    companies = sorted(company_map.values(), key=lambda x: x["name"].lower())
    return {"auto_apply": auto_apply, "manual": manual, "companies": companies}


@app.route("/")
def index():
    jobs = get_jobs(min_score=MIN_FIT_SCORE)
    grouped = _group_jobs(jobs)
    s = stats()
    return render_template(
        "dashboard.html",
        jobs=jobs,
        auto_apply=grouped["auto_apply"],
        manual=grouped["manual"],
        companies=grouped["companies"],
        stats=s,
        threshold=MIN_FIT_SCORE,
        last_updated=datetime.now().strftime("%b %d %Y %I:%M %p"),
    )


@app.route("/api/jobs")
def api_jobs():
    jobs = get_jobs(min_score=MIN_FIT_SCORE)
    return jsonify({"jobs": jobs, "stats": stats(), "apply_state": _apply_state})


@app.route("/api/apply", methods=["POST"])
def api_apply():
    if _apply_state["running"]:
        return jsonify({"error": "already running"}), 409
    data = request.get_json(silent=True) or {}
    job_ids = [int(x) for x in data.get("job_ids", [])]
    if not job_ids:
        return jsonify({"error": "no job_ids"}), 400
    for jid in job_ids:
        update_status(jid, "approved")
    if not _apply_lock.acquire(blocking=False):
        return jsonify({"error": "lock held"}), 409
    try:
        threading.Thread(target=_apply_worker, args=(job_ids,), daemon=True).start()
    finally:
        _apply_lock.release()
    return jsonify({"ok": True, "queued": len(job_ids)})


@app.route("/api/skip", methods=["POST"])
def api_skip():
    data = request.get_json(silent=True) or {}
    job_id = int(data.get("job_id", 0))
    if not job_id:
        return jsonify({"error": "no job_id"}), 400
    update_status(job_id, "skipped")
    return jsonify({"ok": True})


@app.route("/api/mark-applied", methods=["POST"])
def api_mark_applied():
    data = request.get_json(silent=True) or {}
    job_id = int(data.get("job_id", 0))
    if not job_id:
        return jsonify({"error": "no job_id"}), 400
    update_status(job_id, "applied")
    return jsonify({"ok": True})


@app.route("/api/run-now", methods=["POST"])
def api_run_now():
    threading.Thread(target=_search_worker, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/purge-non-us", methods=["POST"])
def api_purge_non_us():
    from db import purge_non_us

    n = purge_non_us()
    return jsonify({"deleted": n})


@app.route("/api/pull-restart", methods=["POST"])
def api_pull_restart():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    try:
        result = subprocess.run(
            ["git", "pull"],
            cwd=project_root,
            capture_output=True,
            text=True,
            timeout=30,
        )
        out = (result.stdout + "\n" + result.stderr).strip()
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"git pull failed: {e}"}), 500

    def _exit_for_relaunch() -> None:
        time.sleep(1)
        os._exit(42)

    threading.Thread(target=_exit_for_relaunch, daemon=True).start()
    return jsonify({"ok": True, "git_output": out, "exiting_with": 42})


@app.route("/status")
def status():
    return jsonify({"apply_state": _apply_state, "stats": stats()})


@app.route("/export.xlsx")
def export_xlsx():
    from openpyxl import Workbook

    jobs = get_jobs(min_score=0)
    grouped = _group_jobs(jobs)

    wb = Workbook()
    # Sheet 1: Auto-Apply
    ws1 = wb.active
    ws1.title = "Auto-Apply"
    _write_jobs_sheet(ws1, grouped["auto_apply"])
    # Sheet 2: Manual
    ws2 = wb.create_sheet("Manual Apply")
    _write_jobs_sheet(ws2, grouped["manual"])
    # Sheet 3: Hiring Manager Contacts
    ws3 = wb.create_sheet("Hiring Managers")
    ws3.append(["Company", "Roles", "LinkedIn search URL", "Contact name (fill in)", "Contact email", "Notes"])
    import urllib.parse as up
    for c in grouped["companies"]:
        q = up.quote(f'"{c["name"]}" ("hiring manager" OR "recruiter" OR "talent acquisition") data')
        url = f"https://www.google.com/search?q=site%3Alinkedin.com%2Fin+{q}"
        ws3.append([c["name"], " · ".join(c["roles"]), url, "", "", ""])
    # Sheet 4: Applied (history)
    applied = [j for j in jobs if j.get("status") == "applied"]
    ws4 = wb.create_sheet("Applied (history)")
    _write_jobs_sheet(ws4, applied)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"job_agent_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _write_jobs_sheet(ws, jobs: list[dict]) -> None:
    ws.append([
        "Title", "Company", "Location", "Work type", "Platform",
        "Fit score", "Status", "Applied at", "Reason", "URL",
    ])
    for j in jobs:
        ws.append([
            j.get("title", ""),
            j.get("company", ""),
            j.get("location", ""),
            j.get("work_type") or "",
            j.get("platform", ""),
            j.get("fit_score", 0),
            j.get("status", ""),
            j.get("applied_at") or "",
            (j.get("fit_reason") or "").replace("\n", " "),
            j.get("url", ""),
        ])


def serve() -> None:
    init_db()
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=False, use_reloader=False)


if __name__ == "__main__":
    serve()
